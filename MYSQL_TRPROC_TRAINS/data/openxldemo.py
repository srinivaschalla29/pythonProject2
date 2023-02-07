import openpyxl
wb=openpyxl.load_workbook("..//data/Book1.xlsx")
print(wb)
sheet=wb.active
data=sheet['A3']
data3=sheet.cell(row=2,column=3).value
data1=sheet['A1:A10']
print(data)
print(data1)
print(data3)
print(sheet.max_row)
print(sheet.min_row)

for i in range(2,8):
    print(sheet.cell(row=i,column=1).value)