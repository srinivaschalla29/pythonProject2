from openpyxl import Workbook
wb=Workbook()
sheet=wb.active
#sheet.tittle="HCL"
#sheet['A1'].value=10
#sheet['B2'].value="test"
#sheet.cell(row=3,column=3).value="Hcl data"
j=0
for i in range(10,60,10):
    j=j+1
    sheet.cell(row=j,column=1).value=i
wb.save("..//data//Book1.xlsx")